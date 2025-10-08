from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, session
from utils import login_required
import os
import re
import io
import csv
from typing import List, Dict
from cache_manager import cached_by_user, invalidate_user_cache
from datetime import datetime

# Se o NET total estiver salvo em centavos no banco, defina NET_TOTAL_IN_CENTS=1 no .env
NET_TOTAL_IN_CENTS = os.getenv("NET_TOTAL_IN_CENTS", "0").lower() in ("1", "true", "yes")

# 🔒 Import protegido: não quebra se Supabase não estiver configurado
try:
    from supabase_client import get_supabase_client
except Exception as e:
    get_supabase_client = None
    import logging
    logging.getLogger(__name__).warning("Supabase indisponível na carga do módulo: %s", e)

def _get_supabase():
    """
    SEGURANÇA: Obtém cliente Supabase autenticado APENAS para o usuário atual.
    Retorna None se não há usuário válido para evitar vazamento de dados.
    """
    if not get_supabase_client:
        return None
    client = get_supabase_client()
    if client is None:
        current_app.logger.debug("CLIENTES: Cliente Supabase não disponível (usuário não autenticado)")
    return client

clientes_bp = Blueprint('clientes', __name__, url_prefix='/clientes')

# ============ ENDPOINT PARA ALOCAR DO MAPA DE LIQUIDEZ ============
@clientes_bp.route('/criar-alocacao-liquidez', methods=['POST'])
@login_required
def criar_alocacao_liquidez():
    """Cria uma alocação a partir do Mapa de Liquidez"""
    from flask import jsonify

    supabase = _get_supabase()
    uid = _uid()

    if not supabase or not uid:
        return jsonify({"success": False, "message": "Sistema indisponível ou sessão inválida"}), 400

    try:
        # Receber dados do formulário
        cliente_codigo = (request.form.get("cliente_codigo") or "").strip()
        produto_id = (request.form.get("produto_id") or "").strip()
        valor = _to_float(request.form.get("valor"))

        if not cliente_codigo or not produto_id:
            return jsonify({"success": False, "message": "Cliente e produto são obrigatórios"}), 400

        if valor <= 0:
            return jsonify({"success": False, "message": "Valor deve ser maior que zero"}), 400

        # Buscar o cliente_id a partir do código
        res_cliente = supabase.table("clientes")\
            .select("id")\
            .eq("user_id", uid)\
            .or_(f"codigo_xp.eq.{cliente_codigo},codigo_mb.eq.{cliente_codigo}")\
            .limit(1)\
            .execute()

        if not res_cliente.data:
            return jsonify({"success": False, "message": "Cliente não encontrado"}), 404

        cliente_id = res_cliente.data[0]["id"]

        # Verificar se já existe alocação para este cliente + produto
        existing_check = supabase.table("alocacoes")\
            .select("id, valor")\
            .eq("cliente_id", cliente_id)\
            .eq("produto_id", produto_id)\
            .eq("user_id", uid)\
            .execute()

        if existing_check.data:
            # Já existe: somar valor
            existing_id = existing_check.data[0]["id"]
            existing_valor = _to_float(existing_check.data[0].get("valor", 0))
            novo_valor = existing_valor + valor

            supabase.table("alocacoes").update({
                "valor": novo_valor
            }).eq("id", existing_id).eq("user_id", uid).execute()

            message = f"Valor adicionado à alocação existente! Novo total: R$ {novo_valor:,.2f}"
        else:
            # Criar nova alocação
            supabase.table("alocacoes").insert({
                "cliente_id": cliente_id,
                "produto_id": produto_id,
                "valor": valor,
                "percentual": 0,
                "efetivada": False,
                "user_id": uid,
            }).execute()

            message = "Alocação criada com sucesso!"

        # Invalidar caches
        from cache_manager import invalidate_all_user_cache
        invalidate_all_user_cache()

        return jsonify({"success": True, "message": message})

    except Exception as e:
        current_app.logger.exception("Erro ao criar alocação do Mapa de Liquidez")
        return jsonify({"success": False, "message": f"Erro ao criar alocação: {str(e)}"}), 500

# ----------------- Helpers -----------------
# ✅ inclui ASSET como modelo permitido
ALLOWED_MODELOS = {"TRADICIONAL", "ASSET", "FEE_BASED", "FEE_BASED_SEM_RV"}

# Headers esperados no template
_TEMPLATE_HEADERS = [
    "nome", "modelo", "repasse",
    "codigo_xp", "codigo_mb",
    "net_xp", "net_xp_global", "net_mb"
]

def _norm_modelo(v: str) -> str:
    """
    Normaliza o texto do modelo para os valores aceitos e
    retorna 'TRADICIONAL' quando vier algo inválido.
    """
    if not v:
        return "TRADICIONAL"
    s = str(v).strip().upper().replace(" ", "_")
    return s if s in ALLOWED_MODELOS else "TRADICIONAL"

def _to_float(x):
    """Converte qualquer coisa para float de forma segura (aceita '1.234,56')."""
    if x is None:
        return 0.0
    s = str(x).strip()
    if s == "" or s.upper() == "NULL":
        return 0.0
    # Remove separador de milhar e normaliza decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        # remove prefixos como R$, % etc.
        m = re.search(r"-?\d+(\.\d+)?", s)
        if m:
            s = m.group(0)
        return float(s)
    except Exception:
        current_app.logger.warning("Valor float inválido: %r", x)
        return 0.0

def _to_int_repasse(x, default=35):
    """
    Converte para int de forma segura; se vier sujo (UUID etc.), cai no default.
    Só permite 35 ou 50.
    """
    if x is None:
        return default
    s = str(x).strip()
    m = re.search(r'-?\d+', s)
    if not m:
        return default
    try:
        v = int(m.group(0))
        return 50 if v == 50 else 35
    except Exception:
        return default

# ----- Filtro Jinja para moeda BRL -----
def _brl(value):
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        return "R$ 0,00"
    s = f"{n:,.2f}"                  # 1234567.89 -> '1,234,567.89'
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")  # -> '1.234.567,89'
    return f"R$ {s}"

clientes_bp.add_app_template_filter(_brl, name="brl")

def _uid():
    # Usar a mesma lógica do security_middleware
    from security_middleware import get_current_user_id
    return get_current_user_id()

# ----------------- Views -----------------
@clientes_bp.route('/')
@login_required
def index():
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return render_template('clientes/index.html',
                             clientes=[],
                             filtros={"q": "", "modelo": "", "letra": ""},
                             stats_by_model={},
                             modelos_ordenados=[],
                             letras_ordenadas=[],
                             total_clientes=0)

    # ── filtros vindos da URL ───────────────────────────────────────────────
    q_txt = (request.args.get('q') or '').strip()              # código XP/MB
    modelo_raw = (request.args.get('modelo') or '').strip()    # modelo
    letra_filter = (request.args.get('letra') or '').strip()   # primeira letra do nome
    modelo_filter = _norm_modelo(modelo_raw) if modelo_raw else ""

    uid = _uid()

    try:
        # monta a query base
        query = supabase.table("clientes").select(
            "id, nome, modelo, repasse, net_total, net_xp, net_xp_global, net_mb, codigo_xp, codigo_mb"
        )
        if uid:
            query = query.eq("user_id", uid)

        # aplica filtro por modelo, se houver
        if modelo_filter:
            query = query.eq("modelo", modelo_filter)

        rows = []
        if q_txt:
            pattern = f"%{q_txt}%"
            # tenta o filtro no banco (OR em codigo_xp OU codigo_mb)
            try:
                query2 = query.or_(f"codigo_xp.ilike.{pattern},codigo_mb.ilike.{pattern}")
                res = query2.execute()
                rows = res.data or []
            except Exception:
                # fallback: busca sem filtro e filtra em memória
                res = query.execute()
                base = res.data or []
                ql = q_txt.lower()
                rows = [
                    r for r in base
                    if ql in str(r.get("codigo_xp") or "").lower()
                    or ql in str(r.get("codigo_mb") or "").lower()
                ]
        else:
            # sem filtro de código
            res = query.execute()
            rows = res.data or []

        # aplica filtro por primeira letra do nome, se houver
        if letra_filter and letra_filter.upper() in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            letra_upper = letra_filter.upper()
            rows = [
                r for r in rows
                if (r.get("nome") or "").strip().upper().startswith(letra_upper)
            ]

        # mapeia/normaliza para o template
        clientes = []
        for r in rows:
            clientes.append({
                "id": r.get("id"),  # UUID do Supabase
                "nome": r.get("nome") or "",
                "modelo": _norm_modelo(r.get("modelo")),
                "repasse": _to_int_repasse(r.get("repasse"), 35),
                # divide por 100 se vier em centavos
                "net_total": (_to_float(r.get("net_total")) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(r.get("net_total")),
                "codigo_xp": (r.get("codigo_xp") or "").strip(),
                "codigo_mb": (r.get("codigo_mb") or "").strip(),
            })

        clientes.sort(key=lambda x: (x["nome"] or "").upper())

        # Calcular estatísticas por modelo
        stats_by_model = {}
        total_clientes = len(clientes)

        for cliente in clientes:
            modelo = cliente.get("modelo", "N/A")
            if modelo not in stats_by_model:
                stats_by_model[modelo] = 0
            stats_by_model[modelo] += 1

        # Ordenar modelos para exibição consistente
        modelos_ordenados = sorted(stats_by_model.keys())

        # Gerar lista de letras disponíveis (buscar todas as primeiras letras dos clientes)
        try:
            all_res = supabase.table("clientes").select("nome").eq("user_id", uid).execute() if uid else None
            all_clientes = all_res.data or [] if all_res else []
            letras_disponiveis = set()
            for c in all_clientes:
                nome = (c.get("nome") or "").strip()
                if nome:
                    primeira_letra = nome[0].upper()
                    if primeira_letra.isalpha():
                        letras_disponiveis.add(primeira_letra)
            letras_ordenadas = sorted(list(letras_disponiveis))
        except Exception:
            letras_ordenadas = []

        return render_template('clientes/index.html',
                             clientes=clientes,
                             filtros={"q": q_txt, "modelo": modelo_filter, "letra": letra_filter},
                             stats_by_model=stats_by_model,
                             modelos_ordenados=modelos_ordenados,
                             letras_ordenadas=letras_ordenadas,
                             total_clientes=total_clientes)

    except Exception:
        current_app.logger.exception("Falha ao listar clientes do Supabase")
        flash("Falha ao listar clientes do Supabase.", "warning")
        return render_template('clientes/index.html',
                             clientes=[],
                             filtros={"q": q_txt, "modelo": modelo_filter, "letra": letra_filter},
                             stats_by_model={},
                             modelos_ordenados=[],
                             letras_ordenadas=[],
                             total_clientes=0)

@clientes_bp.route('/novo', methods=['GET', 'POST'])
@login_required
def novo():
    if request.method == 'POST':
        nome = request.form.get('nome') or ''
        modelo = _norm_modelo(request.form.get('modelo') or 'TRADICIONAL')
        repasse = _to_int_repasse(request.form.get('repasse'), 35)

        cxp = (request.form.get('codigo_xp') or '').strip()
        cmb = (request.form.get('codigo_mb') or '').strip()

        net_xp        = _to_float(request.form.get('net_xp'))
        net_xp_global = _to_float(request.form.get('net_xp_global'))
        net_mb        = _to_float(request.form.get('net_mb'))
        net_total     = net_xp + net_xp_global + net_mb

        supabase = _get_supabase()
        if not supabase:
            flash("Supabase indisponível.", "warning")
            return redirect(url_for('clientes.index'))

        try:
            uid = _uid()
            if not uid:
                flash("Sessão inválida: não foi possível identificar o usuário.", "error")
                return redirect(url_for('clientes.novo'))

            payload = {
                "nome": nome,
                "modelo": modelo,
                "repasse": repasse,
                "codigo_xp": cxp or None,
                "codigo_mb": cmb or None,
                "net_xp": net_xp,
                "net_xp_global": net_xp_global,
                "net_mb": net_mb,
                "net_total": net_total,
                "user_id": uid,  # 👈 garante o dono
            }

            res = supabase.table("clientes").insert(payload).execute()
            current_app.logger.info("Supabase insert clientes -> %s", getattr(res, "data", None))

            # Invalidar cache de clientes
            invalidate_user_cache('clientes_list')

            flash('Cliente cadastrado.', 'success')
        except Exception:
            current_app.logger.exception("Falha ao inserir no Supabase")
            flash('Falha ao inserir no Supabase.', 'warning')

        return redirect(url_for('clientes.index'))

    return render_template('clientes/novo.html')

@clientes_bp.route('/<string:id>/editar', methods=['GET', 'POST'])
@login_required
def editar(id: str):
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return redirect(url_for('clientes.index'))

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return redirect(url_for('clientes.index'))

    if request.method == 'POST':
        nome   = request.form.get('nome') or ''
        modelo = _norm_modelo(request.form.get('modelo') or 'TRADICIONAL')
        repasse = _to_int_repasse(request.form.get('repasse'), 35)

        cxp = (request.form.get('codigo_xp') or '').strip()
        cmb = (request.form.get('codigo_mb') or '').strip()

        net_xp        = _to_float(request.form.get('net_xp'))
        net_xp_global = _to_float(request.form.get('net_xp_global'))
        net_mb        = _to_float(request.form.get('net_mb'))
        net_total     = net_xp + net_xp_global + net_mb

        try:
            payload = {
                "nome": nome,
                "modelo": modelo,
                "repasse": repasse,
                "codigo_xp": cxp or None,
                "codigo_mb": cmb or None,
                "net_xp": net_xp,
                "net_xp_global": net_xp_global,
                "net_mb": net_mb,
                "net_total": net_total,
            }
            # filtra por id e user_id (defensivo se service role)
            supabase.table("clientes").update(payload).eq("id", id).eq("user_id", uid).execute()

            # Invalidar caches relacionados
            invalidate_user_cache('clientes_list')
            invalidate_user_cache('dashboard_data')

            flash('Cliente atualizado com sucesso.', 'success')
        except Exception:
            current_app.logger.exception("Falha ao atualizar no Supabase")
            flash('Falha ao atualizar no Supabase.', 'warning')

        return redirect(url_for('clientes.index'))

    # GET -> carrega o registro para preencher o formulário (filtrando por dono)
    try:
        res = (
            supabase.table("clientes")
            .select("id, nome, modelo, repasse, codigo_xp, codigo_mb, net_xp, net_xp_global, net_mb, net_total")
            .eq("id", id)
            .eq("user_id", uid)
            .limit(1)
            .execute()
        )

        rows = res.data or []
        if not rows:
            flash('Cliente não encontrado no Supabase.', 'warning')
            return redirect(url_for('clientes.index'))

        r = rows[0]
        c = {
            "id": r.get("id"),
            "nome": r.get("nome") or "",
            "modelo": _norm_modelo(r.get("modelo")),
            "repasse": _to_int_repasse(r.get("repasse"), 35),
            "codigo_xp": (r.get("codigo_xp") or ""),
            "codigo_mb": (r.get("codigo_mb") or ""),
            "net_xp": _to_float(r.get("net_xp")),
            "net_xp_global": _to_float(r.get("net_xp_global")),
            "net_mb": _to_float(r.get("net_mb")),
            "_raw_net_total": r.get("net_total"),
            "net_total": (_to_float(r.get("net_total")) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(r.get("net_total")),
        }

        return render_template('clientes/editar.html', c=c)

    except Exception:
        current_app.logger.exception("Falha ao carregar cliente do Supabase")
        flash('Falha ao carregar cliente do Supabase.', 'warning')
        return redirect(url_for('clientes.index'))

@clientes_bp.route('/<string:id>/excluir', methods=['POST'])
@login_required
def excluir(id: str):
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return redirect(url_for('clientes.index'))

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return redirect(url_for('clientes.index'))

    try:
        # apaga apenas o registro do dono
        supabase.table("clientes").delete().eq("id", id).eq("user_id", uid).execute()

        # Invalidar caches relacionados
        invalidate_user_cache('clientes_list')
        invalidate_user_cache('dashboard_data')

        flash('Cliente excluído com sucesso.', 'success')
    except Exception:
        current_app.logger.exception("Falha ao excluir no Supabase")
        flash('Falha ao excluir no Supabase.', 'warning')

    return redirect(url_for('clientes.index'))

# ----------------- Importação em massa -----------------
def _make_template_csv() -> str:
    """Gera um CSV em memória com cabeçalho e 1 linha exemplo."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_TEMPLATE_HEADERS)
    # exemplo (pode editar/duplicar à vontade)
    w.writerow(["Fulano da Silva", "TRADICIONAL", "35", "1234567", "", "1000,00", "0", "500,50"])
    return buf.getvalue()

def _make_template_xlsx() -> bytes:
    """Gera um XLSX em memória com cabeçalho e 1 linha exemplo."""
    try:
        from openpyxl import Workbook
    except Exception:
        raise RuntimeError("Dependência 'openpyxl' não encontrada. Instale-a para baixar o template .xlsx.")
    wb = Workbook()
    ws = wb.active
    ws.title = "ImportarClientes"
    ws.append(_TEMPLATE_HEADERS)
    ws.append(["Fulano da Silva", "TRADICIONAL", "35", "1234567", "", "1000,00", "0", "500,50"])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

def _norm_header(h: str) -> str:
    if not h:
        return ""
    s = h.strip().lower().replace(" ", "_")
    # remoção simples de acentos/c-cedilha
    s = (s.replace("ç","c").replace("ã","a").replace("â","a").replace("á","a")
           .replace("é","e").replace("ê","e").replace("í","i")
           .replace("ó","o").replace("ô","o").replace("ú","u"))
    return s

# mapeamento de sinônimos -> nome interno
_HEADER_MAP = {
    "nome": "nome",
    "modelo": "modelo",
    "repasse": "repasse",
    "codigo_xp": "codigo_xp",
    "codigo_xp.": "codigo_xp",
    "codigo xp": "codigo_xp",
    "código_xp": "codigo_xp",
    "código xp": "codigo_xp",
    "codigo_mb": "codigo_mb",
    "codigo mb": "codigo_mb",
    "código_mb": "codigo_mb",
    "código mb": "codigo_mb",
    "net_xp": "net_xp",
    "net_xp_global": "net_xp_global",
    "net_mb": "net_mb",
    "net total": "net_total",  # ignoramos; derivamos dos outros
}

def _parse_import_file(file_storage) -> List[Dict]:
    """
    Lê .csv ou .xlsx e retorna linhas limpas para 'clientes'.
    - CSV: detecta delimitador (',' ou ';'), mapeia cabeçalhos equivalentes.
    - XLSX: lê a primeira planilha, usa a primeira linha como cabeçalho.
    Converte números com ,/.; normaliza 'modelo' e 'repasse'.
    """
    filename = (getattr(file_storage, "filename", "") or "").lower()

    # --- CSV ---
    if filename.endswith(".csv"):
        content = file_storage.read()
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        # detectar delimitador automaticamente
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ","

        # ler cabeçalho bruto
        buf = io.StringIO(text)
        raw_reader = csv.reader(buf, delimiter=delimiter)
        try:
            raw_headers = next(raw_reader)
        except StopIteration:
            raise ValueError("CSV vazio.")

        headers_norm = [_norm_header(h) for h in raw_headers]
        headers_final = [_HEADER_MAP.get(h, h) for h in headers_norm]

        required = {"nome","modelo","repasse","codigo_xp","codigo_mb","net_xp","net_xp_global","net_mb"}
        present = set(headers_final)
        missing = [h for h in required if h not in present]
        if missing:
            raise ValueError("Template inválido. Faltam colunas: " + ", ".join(sorted(missing)))

        # constrói DictReader com cabeçalhos finais
        buf2 = io.StringIO(text)
        reader = csv.DictReader(buf2, delimiter=delimiter)
        reader.fieldnames = headers_final

        rows = []
        for i, row in enumerate(reader, start=2):
            def g(k): return (row.get(k) or "").strip()
            nome = g("nome")
            if not nome:
                continue
            modelo = _norm_modelo(g("modelo"))
            rep    = _to_int_repasse(g("repasse"), 35)
            cxp = g("codigo_xp")
            cmb = g("codigo_mb")
            net_xp        = _to_float(g("net_xp"))
            net_xp_global = _to_float(g("net_xp_global"))
            net_mb        = _to_float(g("net_mb"))
            net_total     = net_xp + net_xp_global + net_mb

            rows.append({
                "nome": nome,
                "modelo": modelo,
                "repasse": rep,
                "codigo_xp": cxp or None,
                "codigo_mb": cmb or None,
                "net_xp": net_xp,
                "net_xp_global": net_xp_global,
                "net_mb": net_mb,
                "net_total": net_total,
            })
        return rows

    # --- XLSX ---
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise RuntimeError("Dependência 'openpyxl' não encontrada. Instale-a para importar .xlsx.")

        wb = load_workbook(file_storage, data_only=True)
        ws = wb.active  # primeira planilha

        # ler cabeçalhos
        headers_raw = []
        for cell in ws[1]:
            headers_raw.append(str(cell.value or "").strip())
        headers_norm = [_norm_header(h) for h in headers_raw]
        headers_final = [_HEADER_MAP.get(h, h) for h in headers_norm]

        required = {"nome","modelo","repasse","codigo_xp","codigo_mb","net_xp","net_xp_global","net_mb"}
        present = set(headers_final)
        missing = [h for h in required if h not in present]
        if missing:
            raise ValueError("Template inválido. Faltam colunas: " + ", ".join(sorted(missing)))

        # índice por nome final
        idx = {h: i for i, h in enumerate(headers_final)}

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            def gv(key):
                j = idx.get(key)
                return "" if j is None else ("" if r[j] is None else str(r[j]).strip())

            nome = gv("nome")
            if not nome:
                continue
            modelo = _norm_modelo(gv("modelo"))
            rep    = _to_int_repasse(gv("repasse"), 35)
            cxp = gv("codigo_xp")
            cmb = gv("codigo_mb")
            net_xp        = _to_float(gv("net_xp"))
            net_xp_global = _to_float(gv("net_xp_global"))
            net_mb        = _to_float(gv("net_mb"))
            net_total     = net_xp + net_xp_global + net_mb

            rows.append({
                "nome": nome,
                "modelo": modelo,
                "repasse": rep,
                "codigo_xp": cxp or None,
                "codigo_mb": cmb or None,
                "net_xp": net_xp,
                "net_xp_global": net_xp_global,
                "net_mb": net_mb,
                "net_total": net_total,
            })
        return rows

    raise ValueError("Formato não suportado. Envie .csv ou .xlsx.")

@clientes_bp.route('/importar/template.csv', methods=['GET'])
@login_required
def baixar_template_clientes():
    """Baixa um CSV de modelo para importação em massa."""
    csv_text = _make_template_csv()
    return current_app.response_class(
        csv_text,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=template_importar_clientes.csv"
        },
    )

@clientes_bp.route('/importar/template.xlsx', methods=['GET'])
@login_required
def baixar_template_clientes_xlsx():
    """Baixa um XLSX de modelo para importação em massa."""
    xlsx_bytes = _make_template_xlsx()
    return current_app.response_class(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=template_importar_clientes.xlsx"
        },
    )

@clientes_bp.route('/importar', methods=['GET', 'POST'])
@login_required
def importar_em_massa():
    """
    GET  -> mostra a tela com botão de download do(s) template(s) e o upload do arquivo
    POST -> processa o arquivo (.csv ou .xlsx) e insere/upserta clientes do usuário
    """
    if request.method == "GET":
        return render_template("clientes/importar.html")

    # POST
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return redirect(url_for('clientes.index'))

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return redirect(url_for('clientes.index'))

    f = request.files.get("arquivo")
    if not f or not (f.filename.lower().endswith(".csv") or f.filename.lower().endswith(".xlsx")):
        flash("Envie um arquivo .csv ou .xlsx no formato do template.", "error")
        return redirect(url_for("clientes.importar_em_massa"))

    try:
        rows = _parse_import_file(f)
        if not rows:
            flash("Nenhuma linha válida encontrada no arquivo.", "warning")
            return redirect(url_for("clientes.importar_em_massa"))

        # injeta user_id em todas as linhas
        rows = [{**r, "user_id": uid} for r in rows]

        # Estratégia:
        # - preferir UPSERT por (user_id, codigo_xp) quando houver codigo_xp
        # - e por (user_id, codigo_mb) quando houver apenas codigo_mb
        # observação: indices únicos (recomendados) permitem o on_conflict
        #   create unique index if not exists clientes_uid_cxp_uidx on public.clientes (user_id, codigo_xp);
        #   create unique index if not exists clientes_uid_cmb_uidx on public.clientes (user_id, codigo_mb);
        batch_xp = [r for r in rows if r.get("codigo_xp")]
        batch_mb = [r for r in rows if not r.get("codigo_xp") and r.get("codigo_mb")]
        batch_nk = [r for r in rows if not r.get("codigo_xp") and not r.get("codigo_mb")]  # sem chave -> INSERT

        CHUNK = 500

        # 1) upsert por (user_id, codigo_xp)
        for i in range(0, len(batch_xp), CHUNK):
            supabase.table("clientes").upsert(
                batch_xp[i:i+CHUNK],
                on_conflict="user_id,codigo_xp"
            ).execute()

        # 2) upsert por (user_id, codigo_mb)
        for i in range(0, len(batch_mb), CHUNK):
            supabase.table("clientes").upsert(
                batch_mb[i:i+CHUNK],
                on_conflict="user_id,codigo_mb"
            ).execute()

        # 3) inserts simples (sem código de chave) — evitamos sobrescrever nomes aleatoriamente
        for i in range(0, len(batch_nk), CHUNK):
            supabase.table("clientes").insert(batch_nk[i:i+CHUNK]).execute()

        flash(f"Importação concluída. Linhas processadas: {len(rows)}.", "success")
        return redirect(url_for("clientes.index"))

    except ValueError as ve:
        current_app.logger.info("Import clientes: arquivo inválido: %s", ve)
        flash(str(ve), "error")
    except Exception as e:
        current_app.logger.exception("Falha na importação de clientes: %s", e)
        flash("Falha ao importar clientes. Verifique o arquivo e tente novamente.", "error")

    return redirect(url_for("clientes.importar_em_massa"))

@clientes_bp.route('/supernova')
@login_required
def supernova():
    """Tela Supernova: lista clientes com nome e data da última supernova"""
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return render_template('clientes/supernova.html',
                             clientes_supernova=[],
                             filtros={"letra": ""},
                             letras_ordenadas=[])

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return render_template('clientes/supernova.html',
                             clientes_supernova=[],
                             filtros={"letra": ""},
                             letras_ordenadas=[])

    # ── filtros vindos da URL ──────────────────────────────────────────────
    letra_filter = (request.args.get('letra') or '').strip()        # primeira letra do nome
    urgencia_filter = (request.args.get('urgencia') or '').strip()  # nível de urgência
    mes_filter = (request.args.get('mes') or '').strip()            # mês da última supernova

    try:
        # Buscar todos os clientes do usuário
        res_clientes = supabase.table("clientes").select("id, nome").eq("user_id", uid).execute()
        clientes = res_clientes.data or []

        # Aplicar filtro por primeira letra do nome, se houver
        if letra_filter and letra_filter.upper() in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            letra_upper = letra_filter.upper()
            clientes = [
                c for c in clientes
                if (c.get("nome") or "").strip().upper().startswith(letra_upper)
            ]

        # Gerar lista de letras disponíveis (buscar todas as primeiras letras dos clientes)
        try:
            all_res = supabase.table("clientes").select("nome").eq("user_id", uid).execute()
            all_clientes = all_res.data or []
            letras_disponiveis = set()
            for c in all_clientes:
                nome = (c.get("nome") or "").strip()
                if nome:
                    primeira_letra = nome[0].upper()
                    if primeira_letra.isalpha():
                        letras_disponiveis.add(primeira_letra)
            letras_ordenadas = sorted(list(letras_disponiveis))
        except Exception:
            letras_ordenadas = []

        # 🚀 OTIMIZAÇÃO: Buscar TODAS as supernovas de uma vez (elimina N+1)
        supernovas_map = {}
        try:
            res_supernovas_all = (supabase.table("supernovas")
                                .select("cliente_id, data_supernova, observacoes")
                                .eq("user_id", uid)
                                .order("data_supernova", desc=True)
                                .execute())

            # Agrupar por cliente_id, mantendo apenas a mais recente
            for item in (res_supernovas_all.data or []):
                cliente_id = item.get('cliente_id')
                if cliente_id not in supernovas_map:
                    supernovas_map[cliente_id] = item
        except Exception as e:
            current_app.logger.debug("Erro ao buscar supernovas: %s", e)

        clientes_supernova = []
        for cliente in clientes:
            cliente_id = cliente.get('id')
            nome_cliente = cliente.get('nome', '')

            # Buscar do mapa (já carregado)
            supernova_record = supernovas_map.get(cliente_id)

            data_ultima_supernova = None
            observacoes_supernova = None

            if supernova_record:
                data_str = supernova_record.get('data_supernova')
                observacoes_supernova = supernova_record.get('observacoes')

                if data_str:
                    # Converter string para datetime se necessário
                    if isinstance(data_str, str):
                        # Remover timezone se presente e converter
                        data_str_clean = data_str.replace('Z', '').replace('+00:00', '')
                        try:
                            data_ultima_supernova = datetime.fromisoformat(data_str_clean)
                        except ValueError:
                            # Fallback para outros formatos
                            from dateutil import parser
                            data_ultima_supernova = parser.parse(data_str)
                    else:
                        data_ultima_supernova = data_str

            # Calcular urgência baseada na data da última supernova
            urgencia_flag = None
            dias_desde_supernova = None

            if data_ultima_supernova:
                hoje = datetime.now()
                # Se a supernova tem timezone, remover para comparação
                if hasattr(data_ultima_supernova, 'tzinfo') and data_ultima_supernova.tzinfo:
                    data_supernova_local = data_ultima_supernova.replace(tzinfo=None)
                else:
                    data_supernova_local = data_ultima_supernova

                diff = hoje - data_supernova_local
                dias_desde_supernova = diff.days

                if dias_desde_supernova > 90:  # Mais de 3 meses (90 dias)
                    urgencia_flag = 'critica'
                elif dias_desde_supernova > 60:  # Mais de 2 meses
                    urgencia_flag = 'alta'
                elif dias_desde_supernova > 30:  # Mais de 1 mês
                    urgencia_flag = 'media'
                else:
                    urgencia_flag = 'baixa'
            else:
                # Sem supernova = urgência crítica
                urgencia_flag = 'critica'
                dias_desde_supernova = None

            clientes_supernova.append({
                'id': cliente_id,
                'nome': nome_cliente,
                'data_ultima_supernova': data_ultima_supernova,
                'observacoes_supernova': observacoes_supernova,
                'urgencia_flag': urgencia_flag,
                'dias_desde_supernova': dias_desde_supernova
            })

        # Aplicar filtros pós-processamento
        clientes_filtrados = clientes_supernova

        # Filtro por urgência
        if urgencia_filter and urgencia_filter in ['critica', 'alta', 'media', 'baixa']:
            clientes_filtrados = [c for c in clientes_filtrados if c['urgencia_flag'] == urgencia_filter]

        # Filtro por mês da última supernova
        if mes_filter:
            try:
                mes_ano = mes_filter  # Formato esperado: 'YYYY-MM'
                if len(mes_ano) == 7 and '-' in mes_ano:  # Validar formato YYYY-MM
                    clientes_filtrados = [
                        c for c in clientes_filtrados
                        if c['data_ultima_supernova'] and c['data_ultima_supernova'].strftime('%Y-%m') == mes_ano
                    ]
            except Exception:
                current_app.logger.warning("Filtro de mês inválido: %s", mes_filter)

        # Ordenar por urgência (crítica primeiro) e depois por nome
        prioridade_urgencia = {'critica': 0, 'alta': 1, 'media': 2, 'baixa': 3}
        clientes_filtrados.sort(key=lambda x: (
            prioridade_urgencia.get(x['urgencia_flag'], 4),
            (x['nome'] or '').upper()
        ))

        clientes_supernova = clientes_filtrados

        return render_template('clientes/supernova.html',
                             clientes_supernova=clientes_supernova,
                             filtros={
                                 "letra": letra_filter,
                                 "urgencia": urgencia_filter,
                                 "mes": mes_filter
                             },
                             letras_ordenadas=letras_ordenadas)

    except Exception:
        current_app.logger.exception("Falha ao carregar dados da Supernova")
        flash("Falha ao carregar dados da Supernova.", "warning")
        return render_template('clientes/supernova.html',
                             clientes_supernova=[],
                             filtros={
                                 "letra": letra_filter,
                                 "urgencia": urgencia_filter,
                                 "mes": mes_filter
                             },
                             letras_ordenadas=[])

@clientes_bp.route('/supernova/salvar', methods=['POST'])
@login_required
def salvar_supernova():
    """Salva ou atualiza a data da supernova para um cliente"""
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return redirect(url_for('clientes.supernova'))

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return redirect(url_for('clientes.supernova'))

    cliente_id = request.form.get('cliente_id')
    data_supernova_str = request.form.get('data_supernova')
    observacoes = request.form.get('observacoes', '').strip()

    if not cliente_id or not data_supernova_str:
        flash("Cliente e data da supernova são obrigatórios.", "error")
        return redirect(url_for('clientes.supernova'))

    try:
        # Verificar se o cliente pertence ao usuário atual
        res_cliente = (supabase.table("clientes")
                      .select("id, nome")
                      .eq("id", cliente_id)
                      .eq("user_id", uid)
                      .execute())

        if not res_cliente.data:
            flash("Cliente não encontrado ou sem permissão para editar.", "error")
            return redirect(url_for('clientes.supernova'))

        cliente_nome = res_cliente.data[0].get('nome', 'Cliente')

        # Converter string de data para datetime (apenas data, sem hora)
        try:
            # Se vier no formato YYYY-MM-DD (input type="date")
            if len(data_supernova_str) == 10 and '-' in data_supernova_str:
                data_supernova = datetime.strptime(data_supernova_str, '%Y-%m-%d')
            else:
                # Fallback para formatos com hora (compatibilidade)
                data_supernova = datetime.fromisoformat(data_supernova_str)

            current_app.logger.info("SUPERNOVA: Data convertida: %s", data_supernova)
        except ValueError as ve:
            current_app.logger.error("SUPERNOVA: Erro na conversão de data '%s': %s", data_supernova_str, ve)
            flash(f"Formato de data inválido: {data_supernova_str}", "error")
            return redirect(url_for('clientes.supernova'))

        # Primeiro, tentar criar a tabela se não existir
        try:
            # Verificar se já existe um registro de supernova para este cliente
            res_supernova_existente = (supabase.table("supernovas")
                                     .select("id")
                                     .eq("cliente_id", cliente_id)
                                     .eq("user_id", uid)
                                     .execute())

            current_app.logger.info("SUPERNOVA: Verificação existente OK - encontrados %d registros", len(res_supernova_existente.data or []))

        except Exception as table_error:
            current_app.logger.error("SUPERNOVA: Erro ao acessar tabela supernovas: %s", table_error)
            # Se a tabela não existe, vamos tentar criar o registro mesmo assim
            res_supernova_existente = type('MockResponse', (), {'data': []})()

        payload = {
            "cliente_id": cliente_id,
            "data_supernova": data_supernova.isoformat(),
            "observacoes": observacoes or None,
            "user_id": uid,  # Garantir que o registro pertence ao usuário
            "updated_at": datetime.now().isoformat()
        }

        current_app.logger.info("SUPERNOVA: Payload preparado: %s", payload)

        if res_supernova_existente.data:
            # Atualizar registro existente
            supernova_id = res_supernova_existente.data[0]["id"]
            current_app.logger.info("SUPERNOVA: Atualizando registro existente ID: %s", supernova_id)

            result = supabase.table("supernovas").update(payload).eq("id", supernova_id).execute()
            current_app.logger.info("SUPERNOVA: Update result: %s", result.data)

            flash(f"Data da supernova atualizada para {cliente_nome}.", "success")
            current_app.logger.info("SUPERNOVA: Atualizada para cliente %s (%s) by user %s", cliente_nome, cliente_id, uid)
        else:
            # Criar novo registro
            payload["created_at"] = datetime.now().isoformat()
            current_app.logger.info("SUPERNOVA: Criando novo registro para cliente %s", cliente_nome)

            result = supabase.table("supernovas").insert(payload).execute()
            current_app.logger.info("SUPERNOVA: Insert result: %s", result.data)

            flash(f"Data da supernova registrada para {cliente_nome}.", "success")
            current_app.logger.info("SUPERNOVA: Nova criada para cliente %s (%s) by user %s", cliente_nome, cliente_id, uid)

    except Exception as e:
        current_app.logger.exception("SUPERNOVA: Falha completa ao salvar supernova para cliente %s: %s", cliente_id, e)

        # Tentar dar mais detalhes sobre o erro
        error_msg = str(e)
        if "does not exist" in error_msg.lower():
            flash("Tabela de supernova não configurada no banco de dados. Contate o administrador.", "error")
        elif "permission" in error_msg.lower():
            flash("Sem permissão para salvar dados de supernova.", "error")
        elif "constraint" in error_msg.lower():
            flash("Erro de validação nos dados. Verifique os campos.", "error")
        else:
            flash(f"Erro ao salvar data da supernova: {error_msg}", "error")

    return redirect(url_for('clientes.supernova'))

@clientes_bp.route('/supernova/verificar-tabela')
@login_required
def verificar_tabela_supernova():
    """Rota administrativa para verificar/criar tabela supernovas"""
    supabase = _get_supabase()
    if not supabase:
        return {"erro": "Supabase indisponível"}, 500

    try:
        # Tentar fazer uma consulta simples na tabela
        result = supabase.table("supernovas").select("id").limit(1).execute()
        return {
            "sucesso": True,
            "tabela_existe": True,
            "registros_encontrados": len(result.data or []),
            "mensagem": "Tabela 'supernovas' existe e está acessível"
        }
    except Exception as e:
        error_msg = str(e)
        current_app.logger.error("VERIFICAR TABELA SUPERNOVA: %s", error_msg)

        sql_criar_tabela = """
        -- SQL para criar a tabela supernovas no Supabase
        CREATE TABLE IF NOT EXISTS public.supernovas (
            id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
            cliente_id UUID NOT NULL REFERENCES public.clientes(id) ON DELETE CASCADE,
            user_id UUID NOT NULL,
            data_supernova TIMESTAMP WITH TIME ZONE NOT NULL,
            observacoes TEXT,
            created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
            updated_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
        );

        -- Índices para performance
        CREATE INDEX IF NOT EXISTS idx_supernovas_cliente_id ON public.supernovas(cliente_id);
        CREATE INDEX IF NOT EXISTS idx_supernovas_user_id ON public.supernovas(user_id);
        CREATE INDEX IF NOT EXISTS idx_supernovas_data ON public.supernovas(data_supernova);

        -- RLS (Row Level Security)
        ALTER TABLE public.supernovas ENABLE ROW LEVEL SECURITY;

        -- Política para usuários autenticados (Supabase não suporta IF NOT EXISTS em CREATE POLICY)
        DROP POLICY IF EXISTS "Users can manage their own supernovas" ON public.supernovas;
        CREATE POLICY "Users can manage their own supernovas"
        ON public.supernovas
        FOR ALL
        USING (auth.uid() = user_id::uuid);
        """

        return {
            "sucesso": False,
            "tabela_existe": False,
            "erro": error_msg,
            "sql_criar_tabela": sql_criar_tabela,
            "mensagem": "Tabela 'supernovas' não existe. Execute o SQL fornecido no Supabase."
        }, 400

# ----------------- Cross Sell -----------------
@clientes_bp.route('/cross-sell')
@login_required
def cross_sell():
    """Tela Cross Sell: gerencia apresentação e boletagem de produtos aos clientes"""
    import time
    start_time = time.time()

    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return render_template('clientes/cross_sell.html',
                             clientes_cross_sell=[],
                             filtros={"q": "", "letra": ""},
                             letras_ordenadas=[])

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return render_template('clientes/cross_sell.html',
                             clientes_cross_sell=[],
                             filtros={"q": "", "letra": ""},
                             letras_ordenadas=[])

    # Filtros
    q_txt = (request.args.get('q') or '').strip()
    letra_filter = (request.args.get('letra') or '').strip()

    try:
        t1 = time.time()
        current_app.logger.info("CROSS_SELL: Iniciando queries...")
        # Buscar todos os clientes do usuário
        query = supabase.table("clientes").select("id, nome, net_total").eq("user_id", uid)

        if q_txt:
            pattern = f"%{q_txt}%"
            try:
                query = query.ilike("nome", pattern)
            except:
                pass

        res_clientes = query.execute()
        clientes = res_clientes.data or []
        current_app.logger.info("CROSS_SELL: Clientes buscados em %.2fs (%d clientes)", time.time() - t1, len(clientes))

        # Aplicar filtro por primeira letra
        if letra_filter and letra_filter.upper() in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            letra_upper = letra_filter.upper()
            clientes = [
                c for c in clientes
                if (c.get("nome") or "").strip().upper().startswith(letra_upper)
            ]

        # Buscar TODOS os dados de cross sell de uma vez (otimização N+1)
        t2 = time.time()
        try:
            res_cross_all = supabase.table("cross_sell").select("*").eq("user_id", uid).execute()
            cross_sell_map = {item['cliente_id']: item for item in (res_cross_all.data or [])}
            current_app.logger.info("CROSS_SELL: Cross sell data buscado em %.2fs (%d registros)", time.time() - t2, len(cross_sell_map))
        except Exception as e:
            current_app.logger.debug("Erro ao buscar cross_sell: %s", e)
            cross_sell_map = {}

        # Montar lista de clientes com dados de cross sell
        clientes_cross_sell = []
        default_cross_data = {
            'fee_based': '',
            'financial_planning': '',
            'mb': '',
            'offshore': '',
            'produto_estruturado': '',
            'asset': '',
            'seguro_vida': '',
            'consorcio': '',
            'wealth': ''
        }

        for cliente in clientes:
            cliente_id = cliente.get('id')
            nome_cliente = cliente.get('nome', '')

            # Buscar dados do mapa (já carregado)
            cross_data = cross_sell_map.get(cliente_id, default_cross_data)

            clientes_cross_sell.append({
                'id': cliente_id,
                'nome': nome_cliente,
                'fee_based': cross_data.get('fee_based', ''),
                'financial_planning': cross_data.get('financial_planning', ''),
                'mb': cross_data.get('mb', ''),
                'offshore': cross_data.get('offshore', ''),
                'produto_estruturado': cross_data.get('produto_estruturado', ''),
                'asset': cross_data.get('asset', ''),
                'seguro_vida': cross_data.get('seguro_vida', ''),
                'consorcio': cross_data.get('consorcio', ''),
                'wealth': cross_data.get('wealth', '')
            })

        # Ordenar por nome
        clientes_cross_sell.sort(key=lambda x: (x['nome'] or '').upper())

        # Gerar lista de letras disponíveis (usar clientes já carregados)
        letras_disponiveis = set()
        for c in clientes:
            nome = (c.get("nome") or "").strip()
            if nome:
                primeira_letra = nome[0].upper()
                if primeira_letra.isalpha():
                    letras_disponiveis.add(primeira_letra)
        letras_ordenadas = sorted(list(letras_disponiveis))

        elapsed = time.time() - start_time
        current_app.logger.info("CROSS_SELL: Página carregada em %.2fs", elapsed)

        return render_template('clientes/cross_sell.html',
                             clientes_cross_sell=clientes_cross_sell,
                             filtros={"q": q_txt, "letra": letra_filter},
                             letras_ordenadas=letras_ordenadas)

    except Exception:
        current_app.logger.exception("Falha ao carregar dados de Cross Sell")
        flash("Falha ao carregar dados de Cross Sell.", "warning")
        return render_template('clientes/cross_sell.html',
                             clientes_cross_sell=[],
                             filtros={"q": q_txt, "letra": letra_filter},
                             letras_ordenadas=[])

@clientes_bp.route('/cross-sell/salvar', methods=['POST'])
@login_required
def salvar_cross_sell():
    """Salva ou atualiza os dados de cross sell para um cliente"""
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return redirect(url_for('clientes.cross_sell'))

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return redirect(url_for('clientes.cross_sell'))

    cliente_id = request.form.get('cliente_id')

    if not cliente_id:
        flash("Cliente não identificado.", "error")
        return redirect(url_for('clientes.cross_sell'))

    try:
        # Verificar se o cliente pertence ao usuário atual
        res_cliente = (supabase.table("clientes")
                      .select("id, nome")
                      .eq("id", cliente_id)
                      .eq("user_id", uid)
                      .execute())

        if not res_cliente.data:
            flash("Cliente não encontrado ou sem permissão para editar.", "error")
            return redirect(url_for('clientes.cross_sell'))

        cliente_nome = res_cliente.data[0].get('nome', 'Cliente')

        # Coletar dados do formulário
        payload = {
            "cliente_id": cliente_id,
            "user_id": uid,
            "fee_based": request.form.get('fee_based', ''),
            "financial_planning": request.form.get('financial_planning', ''),
            "mb": request.form.get('mb', ''),
            "offshore": request.form.get('offshore', ''),
            "produto_estruturado": request.form.get('produto_estruturado', ''),
            "asset": request.form.get('asset', ''),
            "seguro_vida": request.form.get('seguro_vida', ''),
            "consorcio": request.form.get('consorcio', ''),
            "wealth": request.form.get('wealth', ''),
            "updated_at": datetime.now().isoformat()
        }

        # Verificar se já existe um registro
        res_existente = (supabase.table("cross_sell")
                        .select("id")
                        .eq("cliente_id", cliente_id)
                        .eq("user_id", uid)
                        .execute())

        if res_existente.data:
            # Atualizar
            cross_sell_id = res_existente.data[0]["id"]
            supabase.table("cross_sell").update(payload).eq("id", cross_sell_id).execute()
            flash(f"Cross Sell atualizado para {cliente_nome}.", "success")
        else:
            # Inserir
            payload["created_at"] = datetime.now().isoformat()
            supabase.table("cross_sell").insert(payload).execute()
            flash(f"Cross Sell registrado para {cliente_nome}.", "success")

    except Exception as e:
        current_app.logger.exception("Falha ao salvar cross sell: %s", e)
        flash(f"Erro ao salvar Cross Sell: {str(e)}", "error")

    return redirect(url_for('clientes.cross_sell'))

# ----------------- Insights -----------------
@clientes_bp.route('/insights')
@login_required
def insights():
    """Tela Insights: cruzamento de dados para insights valiosos"""
    import time
    start_time = time.time()

    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return render_template('clientes/insights.html', insights=[])

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return render_template('clientes/insights.html', insights=[])

    try:
        # Buscar todos os dados necessários
        t1 = time.time()
        res_clientes = supabase.table("clientes").select("*").eq("user_id", uid).execute()
        clientes = res_clientes.data or []
        current_app.logger.info("INSIGHTS: Clientes buscados em %.2fs (%d clientes)", time.time() - t1, len(clientes))

        # Buscar dados de cross_sell
        t2 = time.time()
        res_cross = supabase.table("cross_sell").select("*").eq("user_id", uid).execute()
        cross_sell_data = {item['cliente_id']: item for item in (res_cross.data or [])}
        current_app.logger.info("INSIGHTS: Cross sell buscado em %.2fs (%d registros)", time.time() - t2, len(cross_sell_data))

        insights_list = []

        # === INSIGHT 1: Clientes com NET > 1M sem Financial Planning ===
        clientes_alto_net_sem_fp = []
        for cliente in clientes:
            net_total = (_to_float(cliente.get('net_total')) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(cliente.get('net_total'))
            if net_total >= 1000000:
                cross = cross_sell_data.get(cliente['id'], {})
                fp_status = cross.get('financial_planning', '')
                if fp_status != 'Apresentado' and fp_status != 'Boletado':
                    clientes_alto_net_sem_fp.append({
                        'nome': cliente.get('nome'),
                        'net_total': net_total
                    })

        if clientes_alto_net_sem_fp:
            insights_list.append({
                'tipo': 'oportunidade',
                'titulo': f'🎯 {len(clientes_alto_net_sem_fp)} clientes com NET > R$ 1M sem Financial Planning',
                'descricao': f'Há {len(clientes_alto_net_sem_fp)} clientes com patrimônio acima de R$ 1 milhão que ainda não tiveram Financial Planning apresentado.',
                'clientes': clientes_alto_net_sem_fp,
                'prioridade': 'alta'
            })

        # === INSIGHT 2: Clientes TRADICIONAL com NET > 500K (candidatos a Fee Based) ===
        clientes_tradicional_candidatos_fee = []
        for cliente in clientes:
            modelo = _norm_modelo(cliente.get('modelo'))
            net_total = (_to_float(cliente.get('net_total')) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(cliente.get('net_total'))
            if modelo == 'TRADICIONAL' and net_total >= 500000:
                cross = cross_sell_data.get(cliente['id'], {})
                fee_status = cross.get('fee_based', '')
                if fee_status != 'Apresentado' and fee_status != 'Boletado':
                    clientes_tradicional_candidatos_fee.append({
                        'nome': cliente.get('nome'),
                        'net_total': net_total
                    })

        if clientes_tradicional_candidatos_fee:
            insights_list.append({
                'tipo': 'conversao',
                'titulo': f'💡 {len(clientes_tradicional_candidatos_fee)} clientes TRADICIONAL com NET > R$ 500K',
                'descricao': f'{len(clientes_tradicional_candidatos_fee)} clientes no modelo TRADICIONAL com patrimônio acima de R$ 500 mil são fortes candidatos para migração ao modelo Fee Based.',
                'clientes': clientes_tradicional_candidatos_fee,
                'prioridade': 'alta'
            })

        # === INSIGHT 3: Clientes sem Offshore com NET > 1M ===
        clientes_sem_offshore = []
        for cliente in clientes:
            net_total = (_to_float(cliente.get('net_total')) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(cliente.get('net_total'))
            if net_total >= 1000000:
                cross = cross_sell_data.get(cliente['id'], {})
                offshore_status = cross.get('offshore', '')
                if offshore_status != 'Apresentado' and offshore_status != 'Boletado':
                    clientes_sem_offshore.append({
                        'nome': cliente.get('nome'),
                        'net_total': net_total
                    })

        if clientes_sem_offshore:
            insights_list.append({
                'tipo': 'oportunidade',
                'titulo': f'🌎 {len(clientes_sem_offshore)} clientes com NET > R$ 1M sem Offshore',
                'descricao': f'{len(clientes_sem_offshore)} clientes com patrimônio acima de R$ 1 milhão ainda não tiveram soluções Offshore apresentadas.',
                'clientes': clientes_sem_offshore,
                'prioridade': 'alta'
            })

        # === INSIGHT 4: Clientes com NET > 3M sem Asset apresentado ===
        clientes_alto_net_sem_asset = []
        for cliente in clientes:
            net_total = (_to_float(cliente.get('net_total')) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(cliente.get('net_total'))
            if net_total >= 3000000:
                cross = cross_sell_data.get(cliente['id'], {})
                asset_status = cross.get('asset', '')
                if asset_status != 'Apresentado' and asset_status != 'Boletado':
                    clientes_alto_net_sem_asset.append({
                        'nome': cliente.get('nome'),
                        'net_total': net_total
                    })

        if clientes_alto_net_sem_asset:
            insights_list.append({
                'tipo': 'oportunidade',
                'titulo': f'💎 {len(clientes_alto_net_sem_asset)} clientes com NET > R$ 3M sem Asset apresentado',
                'descricao': f'{len(clientes_alto_net_sem_asset)} clientes com patrimônio acima de R$ 3 milhões ainda não tiveram o modelo Asset apresentado.',
                'clientes': clientes_alto_net_sem_asset,
                'prioridade': 'alta'
            })

        # === INSIGHT 5: Clientes sem Seguro de Vida (todos com NET > 300K) ===
        clientes_sem_seguro = []
        for cliente in clientes:
            net_total = (_to_float(cliente.get('net_total')) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(cliente.get('net_total'))
            if net_total >= 300000:
                cross = cross_sell_data.get(cliente['id'], {})
                seguro_status = cross.get('seguro_vida', '')
                if seguro_status != 'Apresentado' and seguro_status != 'Boletado':
                    clientes_sem_seguro.append({
                        'nome': cliente.get('nome'),
                        'net_total': net_total
                    })

        if clientes_sem_seguro:
            insights_list.append({
                'tipo': 'protecao',
                'titulo': f'🛡️ {len(clientes_sem_seguro)} clientes com NET > R$ 300K sem Seguro de Vida',
                'descricao': f'{len(clientes_sem_seguro)} clientes com patrimônio significativo ainda não têm Seguro de Vida apresentado.',
                'clientes': clientes_sem_seguro,
                'prioridade': 'media'
            })

        # === INSIGHT 6: Distribuição de modelos (estatística) ===
        distribuicao_modelos = {}
        total_net_por_modelo = {}
        for cliente in clientes:
            modelo = _norm_modelo(cliente.get('modelo'))
            net_total = (_to_float(cliente.get('net_total')) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(cliente.get('net_total'))
            distribuicao_modelos[modelo] = distribuicao_modelos.get(modelo, 0) + 1
            total_net_por_modelo[modelo] = total_net_por_modelo.get(modelo, 0) + net_total

        modelo_mais_comum = max(distribuicao_modelos.items(), key=lambda x: x[1]) if distribuicao_modelos else None
        if modelo_mais_comum:
            insights_list.append({
                'tipo': 'estatistica',
                'titulo': f'📈 Distribuição de Modelos: {modelo_mais_comum[0]} lidera com {modelo_mais_comum[1]} clientes',
                'descricao': f'Seu modelo mais comum é {modelo_mais_comum[0]} com {modelo_mais_comum[1]} clientes. NET total neste modelo: R$ {total_net_por_modelo.get(modelo_mais_comum[0], 0):,.2f}',
                'clientes': [],
                'prioridade': 'info'
            })

        # === INSIGHT 7: Clientes com MB não apresentado ===
        clientes_sem_mb = []
        for cliente in clientes:
            codigo_mb = (cliente.get('codigo_mb') or '').strip()
            if not codigo_mb:  # Cliente não tem código MB
                cross = cross_sell_data.get(cliente['id'], {})
                mb_status = cross.get('mb', '')
                if mb_status != 'Apresentado' and mb_status != 'Boletado':
                    net_total = (_to_float(cliente.get('net_total')) / 100.0) if NET_TOTAL_IN_CENTS else _to_float(cliente.get('net_total'))
                    if net_total >= 100000:  # Só clientes com NET significativo
                        clientes_sem_mb.append({
                            'nome': cliente.get('nome'),
                            'net_total': net_total
                        })

        if clientes_sem_mb:
            insights_list.append({
                'tipo': 'expansao',
                'titulo': f'🏦 {len(clientes_sem_mb)} clientes sem MB (NET > R$ 100K)',
                'descricao': f'{len(clientes_sem_mb)} clientes com patrimônio acima de R$ 100 mil ainda não têm conta MB. Oportunidade de expansão.',
                'clientes': clientes_sem_mb,
                'prioridade': 'media'
            })

        # Ordenar insights por prioridade
        prioridade_ordem = {'alta': 0, 'media': 1, 'info': 2}
        insights_list.sort(key=lambda x: prioridade_ordem.get(x['prioridade'], 3))

        elapsed = time.time() - start_time
        current_app.logger.info("INSIGHTS: Página carregada em %.2fs (%d insights)", elapsed, len(insights_list))

        return render_template('clientes/insights.html', insights=insights_list)

    except Exception:
        current_app.logger.exception("Falha ao gerar insights")
        flash("Falha ao gerar insights.", "warning")
        return render_template('clientes/insights.html', insights=[])


@clientes_bp.route('/asset-allocation')
@login_required
def asset_allocation():
    """Tela Asset Allocation: visualização da distribuição de ativos por cliente"""
    supabase = _get_supabase()
    if not supabase:
        flash("Supabase indisponível.", "warning")
        return render_template('clientes/asset_allocation.html',
                             clientes=[],
                             cliente_selecionado=None,
                             por_produto=[],
                             por_sub_produto=[],
                             por_ativo=[],
                             por_emissor=[],
                             credito_privado_vencimentos=[],
                             credito_privado_emissores=[],
                             exposicao_rf={})

    uid = _uid()
    if not uid:
        flash("Sessão inválida: não foi possível identificar o usuário.", "error")
        return render_template('clientes/asset_allocation.html',
                             clientes=[],
                             cliente_selecionado=None,
                             por_produto=[],
                             por_sub_produto=[],
                             por_ativo=[],
                             por_emissor=[],
                             credito_privado_vencimentos=[],
                             credito_privado_emissores=[],
                             exposicao_rf={})

    try:
        # Buscar lista de clientes (códigos únicos do diversificador)
        res_clientes = supabase.table("diversificador")\
            .select("cliente")\
            .eq("user_id", uid)\
            .execute()

        # Extrair códigos únicos de clientes
        codigos_clientes = list(set(item['cliente'] for item in (res_clientes.data or []) if item.get('cliente')))
        codigos_clientes.sort()

        # Buscar nomes dos clientes da tabela clientes
        clientes_com_nome = []
        res_nomes = supabase.table("clientes")\
            .select("codigo_xp, codigo_mb, nome")\
            .eq("user_id", uid)\
            .execute()

        # Criar mapeamento código -> nome
        codigo_para_nome = {}
        for cliente in (res_nomes.data or []):
            codigo_xp = (cliente.get('codigo_xp') or '').strip()
            codigo_mb = (cliente.get('codigo_mb') or '').strip()
            nome = cliente.get('nome', '')
            if codigo_xp:
                codigo_para_nome[codigo_xp] = nome
            if codigo_mb:
                codigo_para_nome[codigo_mb] = nome

        # Criar lista de clientes com nome para o dropdown (apenas os encontrados na tabela clientes)
        for codigo in codigos_clientes:
            if codigo in codigo_para_nome:
                clientes_com_nome.append({
                    'codigo': codigo,
                    'nome': codigo_para_nome[codigo]
                })

        # Ordenar por nome
        clientes_com_nome.sort(key=lambda x: x['nome'].upper())

        # Calcular letras iniciais disponíveis
        letras_disponiveis = set()
        for cliente in clientes_com_nome:
            nome = cliente.get('nome', '')
            if nome:
                primeira_letra = nome[0].upper()
                if primeira_letra.isalpha():
                    letras_disponiveis.add(primeira_letra)
        letras_disponiveis = sorted(letras_disponiveis)

        # Cliente selecionado (via query param)
        cliente_selecionado = request.args.get('cliente')
        cliente_selecionado_nome = None

        # Dados de posição do cliente
        por_produto = []
        por_sub_produto = []
        por_ativo = []
        por_emissor = []
        exposicao_por_tipo = {}
        credito_privado_vencimentos = []
        credito_privado_emissores = []
        exposicao_rf = {}
        mapa_liquidez = []

        if cliente_selecionado:
            # Buscar o nome do cliente selecionado
            cliente_selecionado_nome = codigo_para_nome.get(cliente_selecionado, f'Cliente {cliente_selecionado}')
            # Buscar todas as posições do cliente selecionado
            res_posicoes = supabase.table("diversificador")\
                .select("*")\
                .eq("user_id", uid)\
                .eq("cliente", cliente_selecionado)\
                .execute()

            posicoes = res_posicoes.data or []

            # Agrupar por Produto
            agrupamento_produto = {}
            for pos in posicoes:
                produto = pos.get('produto', 'Não Informado')
                net = _to_float(pos.get('net', 0))
                if produto not in agrupamento_produto:
                    agrupamento_produto[produto] = 0.0
                agrupamento_produto[produto] += net

            por_produto = [{'categoria': k, 'valor': v} for k, v in agrupamento_produto.items()]
            por_produto.sort(key=lambda x: x['valor'], reverse=True)

            # Agrupar por Sub Produto
            agrupamento_sub_produto = {}
            for pos in posicoes:
                sub_produto = pos.get('sub_produto', 'Não Informado')
                net = _to_float(pos.get('net', 0))
                if sub_produto not in agrupamento_sub_produto:
                    agrupamento_sub_produto[sub_produto] = 0.0
                agrupamento_sub_produto[sub_produto] += net

            por_sub_produto = [{'categoria': k, 'valor': v} for k, v in agrupamento_sub_produto.items()]
            por_sub_produto.sort(key=lambda x: x['valor'], reverse=True)

            # Agrupar por Ativo (top 20)
            agrupamento_ativo = {}
            for pos in posicoes:
                ativo = pos.get('ativo', 'Não Informado')
                net = _to_float(pos.get('net', 0))
                if ativo not in agrupamento_ativo:
                    agrupamento_ativo[ativo] = 0.0
                agrupamento_ativo[ativo] += net

            por_ativo = [{'categoria': k, 'valor': v} for k, v in agrupamento_ativo.items()]
            por_ativo.sort(key=lambda x: x['valor'], reverse=True)
            por_ativo = por_ativo[:20]  # Top 20 ativos

            # Agrupar por Emissor (top 20)
            agrupamento_emissor = {}
            for pos in posicoes:
                emissor = pos.get('emissor', 'Não Informado')
                net = _to_float(pos.get('net', 0))
                if emissor not in agrupamento_emissor:
                    agrupamento_emissor[emissor] = 0.0
                agrupamento_emissor[emissor] += net

            por_emissor = [{'categoria': k, 'valor': v} for k, v in agrupamento_emissor.items()]
            por_emissor.sort(key=lambda x: x['valor'], reverse=True)
            por_emissor = por_emissor[:20]  # Top 20 emissores

            # === CLASSIFICAÇÃO POR TIPO DE ATIVO ===
            # O sub_produto já vem com a categoria (ex: "Emissão Bancária", "Título Público", "Crédito Privado")
            # Vamos buscar também o campo "produto" para ter detalhes (CDB, LCI, etc)

            # Inicializar contadores
            total_emissao_bancaria = 0.0
            total_credito_privado = 0.0
            total_titulo_publico = 0.0
            total_geral = 0.0

            # Detalhes por tipo dentro de cada categoria
            detalhes_bancaria = {}  # produto -> valor
            detalhes_credito = {}
            detalhes_publico = {}

            current_app.logger.info(f"ASSET_ALLOCATION: Processando {len(posicoes)} posições para cliente {cliente_selecionado}")

            # Primeiro, vamos buscar as posições de RF da tabela custodia_rf para este cliente
            try:
                res_custodia_rf_tipos = supabase.table("custodia_rf")\
                    .select("cod_conta, nome_papel, custodia")\
                    .eq("user_id", uid)\
                    .eq("cod_conta", str(cliente_selecionado))\
                    .execute()

                custodia_rf_dict = {}  # nome_papel -> custodia
                if res_custodia_rf_tipos.data:
                    for item in res_custodia_rf_tipos.data:
                        nome_papel = (item.get('nome_papel', '') or '').strip().upper()
                        custodia = _to_float(item.get('custodia', 0))
                        custodia_rf_dict[nome_papel] = custodia_rf_dict.get(nome_papel, 0.0) + custodia

                    current_app.logger.info(f"ASSET_ALLOCATION: Encontrados {len(custodia_rf_dict)} ativos RF únicos na custodia_rf")
            except Exception as e:
                current_app.logger.warning(f"ASSET_ALLOCATION: Erro ao buscar custodia_rf para classificação de tipos: {e}")
                custodia_rf_dict = {}

            # Agora classificar usando nome_papel
            for nome_papel, valor in custodia_rf_dict.items():
                total_geral += valor

                # Tesouro Direto: nome_papel vazio (NULL)
                if not nome_papel:
                    total_titulo_publico += valor
                    detalhes_publico['Tesouro Direto'] = detalhes_publico.get('Tesouro Direto', 0.0) + valor
                    current_app.logger.debug(f"ASSET_ALLOCATION: Classificado como Título Público (Tesouro Direto): nome_papel vazio -> R$ {valor:,.2f}")

                # Verificar o prefixo do nome_papel
                # Emissão Bancária: CDB, LCD, LCI, LCA, LF
                elif nome_papel.startswith(('CDB', 'LCD', 'LCI', 'LCA', 'LF')):
                    total_emissao_bancaria += valor
                    # Extrair o tipo (primeiras 3 letras geralmente)
                    tipo_detalhe = nome_papel.split()[0] if ' ' in nome_papel else nome_papel[:3]
                    detalhes_bancaria[tipo_detalhe] = detalhes_bancaria.get(tipo_detalhe, 0.0) + valor
                    current_app.logger.debug(f"ASSET_ALLOCATION: Classificado como Emissão Bancária: {nome_papel} -> {tipo_detalhe}")

                # Crédito Privado: DEB, CDCA, CRA, CRI, FIDC
                elif nome_papel.startswith(('DEB', 'CDCA', 'CRA', 'CRI', 'FIDC')):
                    total_credito_privado += valor
                    tipo_detalhe = nome_papel.split()[0] if ' ' in nome_papel else nome_papel[:4] if nome_papel.startswith('CDCA') or nome_papel.startswith('FIDC') else nome_papel[:3]
                    detalhes_credito[tipo_detalhe] = detalhes_credito.get(tipo_detalhe, 0.0) + valor
                    current_app.logger.debug(f"ASSET_ALLOCATION: Classificado como Crédito Privado: {nome_papel} -> {tipo_detalhe}")

                # Título Público: NTN-B, LFT, LTN, NTN-F, NTN-C
                elif nome_papel.startswith(('NTN-B', 'LFT', 'LTN', 'NTN-F', 'NTN-C')):
                    total_titulo_publico += valor
                    tipo_detalhe = nome_papel.split()[0] if ' ' in nome_papel else ('NTN-B' if nome_papel.startswith('NTN-B') else 'NTN-F' if nome_papel.startswith('NTN-F') else 'NTN-C' if nome_papel.startswith('NTN-C') else nome_papel[:3])
                    detalhes_publico[tipo_detalhe] = detalhes_publico.get(tipo_detalhe, 0.0) + valor
                    current_app.logger.debug(f"ASSET_ALLOCATION: Classificado como Título Público: {nome_papel} -> {tipo_detalhe}")

                else:
                    current_app.logger.warning(f"ASSET_ALLOCATION: NÃO CLASSIFICADO: '{nome_papel}'")

            current_app.logger.info(f"ASSET_ALLOCATION: Totais - Bancária: R$ {total_emissao_bancaria:,.2f} | Crédito: R$ {total_credito_privado:,.2f} | Público: R$ {total_titulo_publico:,.2f} | Total Geral: R$ {total_geral:,.2f}")
            current_app.logger.info(f"ASSET_ALLOCATION: Detalhes Bancária: {detalhes_bancaria}")
            current_app.logger.info(f"ASSET_ALLOCATION: Detalhes Crédito: {detalhes_credito}")
            current_app.logger.info(f"ASSET_ALLOCATION: Detalhes Público: {detalhes_publico}")

            # === ANÁLISE DE VENCIMENTOS E EMISSORES (TODOS OS ATIVOS RF) ===
            credito_privado_vencimentos = []
            credito_privado_emissores = []

            # Buscar dados de toda a tabela custodia_rf para vencimentos e emissores
            try:
                res_cp_rf = supabase.table("custodia_rf")\
                    .select("*")\
                    .eq("user_id", uid)\
                    .eq("cod_conta", str(cliente_selecionado))\
                    .execute()

                if res_cp_rf.data:
                    current_app.logger.info(f"ASSET_ALLOCATION: Encontradas {len(res_cp_rf.data)} posições na custodia_rf para análise de vencimentos")

                    # === 1. ANÁLISE POR VENCIMENTO (TODOS OS ATIVOS RF - USAR COLUNA VENCIMENTO) ===
                    vencimentos_por_ano = {}  # {2025: {'total': 0.0, 'count': 0}, ...}
                    total_com_vencimento = 0.0

                    for pos in res_cp_rf.data:
                        custodia = _to_float(pos.get('custodia', 0))
                        vencimento = pos.get('vencimento')  # É uma data no formato 'YYYY-MM-DD' ou objeto date

                        if vencimento:
                            try:
                                # Extrair o ano da data de vencimento
                                # Se for string no formato 'YYYY-MM-DD', pegar os primeiros 4 caracteres
                                # Se for um objeto date, usar .year
                                if isinstance(vencimento, str):
                                    # Formato: '2025-12-31' -> pegar '2025'
                                    ano = int(vencimento[:4])
                                elif hasattr(vencimento, 'year'):
                                    # É um objeto date/datetime
                                    ano = vencimento.year
                                else:
                                    # Tentar converter direto para int (fallback)
                                    ano = int(vencimento)

                                if ano not in vencimentos_por_ano:
                                    vencimentos_por_ano[ano] = {'total': 0.0, 'count': 0}

                                vencimentos_por_ano[ano]['total'] += custodia
                                vencimentos_por_ano[ano]['count'] += 1
                                total_com_vencimento += custodia
                            except (ValueError, TypeError, IndexError) as e:
                                current_app.logger.warning(f"ASSET_ALLOCATION: Erro ao extrair ano do vencimento '{vencimento}' (tipo: {type(vencimento).__name__}): {e}")

                    # Preparar lista de vencimentos para o template (ordenado por ano)
                    for ano in sorted(vencimentos_por_ano.keys()):
                        dados = vencimentos_por_ano[ano]
                        credito_privado_vencimentos.append({
                            'label': str(ano),
                            'total': dados['total'],
                            'count': dados['count'],
                            'percentual_carteira': (dados['total'] / total_com_vencimento * 100) if total_com_vencimento > 0 else 0
                        })

                    current_app.logger.info(f"ASSET_ALLOCATION: Vencimentos por ano: {credito_privado_vencimentos}")
                    current_app.logger.info(f"ASSET_ALLOCATION: Total de RF com vencimento informado: R$ {total_com_vencimento:,.2f}")

                    # === 2. ANÁLISE POR EMISSOR (EXTRAIR DO NOME_PAPEL) ===
                    emissores_agrupados = {}

                    for pos in res_cp_rf.data:
                        nome_papel = (pos.get('nome_papel', '') or '').strip()
                        custodia = _to_float(pos.get('custodia', 0))

                        # Extrair emissor do nome_papel e identificar tipo de ativo
                        emissor = 'Não Informado'
                        tipo_ativo = 'outros'  # Pode ser 'emissao_bancaria', 'credito_privado', 'titulo_publico', ou 'outros'

                        # Se nome_papel estiver vazio, considerar como Tesouro Nacional
                        if not nome_papel:
                            emissor = 'Tesouro Nacional'
                            tipo_ativo = 'titulo_publico'
                        # Verificar se é título público
                        elif any(titulo in nome_papel for titulo in ['NTN-B', 'LFT', 'NTN-C', 'LTN', 'NTN-F']):
                            emissor = 'Tesouro Nacional'
                            tipo_ativo = 'titulo_publico'
                        else:
                            # Identificar tipo de ativo baseado no prefixo
                            nome_upper = nome_papel.upper()
                            if nome_upper.startswith(('CDB', 'LCD', 'LCI', 'LCA', 'LF')):
                                tipo_ativo = 'emissao_bancaria'
                            elif nome_upper.startswith(('DEB', 'CDCA', 'CRA', 'CRI', 'FIDC')):
                                tipo_ativo = 'credito_privado'

                            # Extrair emissor: o que vem após o primeiro espaço e antes do -
                            # Exemplo: "CRA JBS - SET/2032" -> "JBS"
                            partes = nome_papel.split()
                            if len(partes) >= 2:
                                # Pegar tudo após a primeira palavra até encontrar um -
                                resto = ' '.join(partes[1:])
                                if ' - ' in resto:
                                    emissor = resto.split(' - ')[0].strip()
                                elif '-' in resto:
                                    emissor = resto.split('-')[0].strip()
                                else:
                                    emissor = partes[1].strip()

                        if emissor not in emissores_agrupados:
                            emissores_agrupados[emissor] = {'total': 0.0, 'count': 0, 'tipo_ativo': tipo_ativo}

                        emissores_agrupados[emissor]['total'] += custodia
                        emissores_agrupados[emissor]['count'] += 1

                    # Preparar lista de emissores para o template (top 10)
                    for emissor, dados in emissores_agrupados.items():
                        # Só destacar acima do FGC se for Emissão Bancária
                        acima_fgc = (dados['tipo_ativo'] == 'emissao_bancaria' and dados['total'] > 250000)

                        credito_privado_emissores.append({
                            'nome': emissor,
                            'total': dados['total'],
                            'count': dados['count'],
                            'percentual_carteira': (dados['total'] / total_geral * 100) if total_geral > 0 else 0,
                            'acima_fgc': acima_fgc,  # Flag para destacar emissões bancárias acima de 250k
                            'tipo_ativo': dados['tipo_ativo']
                        })

                    # Ordenar por percentual decrescente e pegar top 10
                    credito_privado_emissores.sort(key=lambda x: x['percentual_carteira'], reverse=True)
                    credito_privado_emissores = credito_privado_emissores[:10]

                    current_app.logger.info(f"ASSET_ALLOCATION: Emissores (top 10): {credito_privado_emissores}")

                else:
                    current_app.logger.warning(f"ASSET_ALLOCATION: Nenhuma posição encontrada na custodia_rf para cliente {cliente_selecionado}")

            except Exception as e:
                current_app.logger.warning(f"ASSET_ALLOCATION: Erro ao buscar vencimentos/emissores da custodia_rf: {e}")

            # Calcular total de Renda Fixa (soma das 3 categorias)
            total_renda_fixa = total_emissao_bancaria + total_credito_privado + total_titulo_publico

            # Preparar dados para o template
            exposicao_por_tipo = {
                'emissao_bancaria': {
                    'total': total_emissao_bancaria,
                    'percentual': (total_emissao_bancaria / total_geral * 100) if total_geral > 0 else 0,
                    'percentual_rf': (total_emissao_bancaria / total_renda_fixa * 100) if total_renda_fixa > 0 else 0,
                    'detalhes': [{'tipo': k, 'valor': v, 'percentual': (v / total_emissao_bancaria * 100) if total_emissao_bancaria > 0 else 0}
                                for k, v in sorted(detalhes_bancaria.items(), key=lambda x: x[1], reverse=True)]
                },
                'credito_privado': {
                    'total': total_credito_privado,
                    'percentual': (total_credito_privado / total_geral * 100) if total_geral > 0 else 0,
                    'percentual_rf': (total_credito_privado / total_renda_fixa * 100) if total_renda_fixa > 0 else 0,
                    'detalhes': [{'tipo': k, 'valor': v, 'percentual': (v / total_credito_privado * 100) if total_credito_privado > 0 else 0}
                                for k, v in sorted(detalhes_credito.items(), key=lambda x: x[1], reverse=True)]
                },
                'titulo_publico': {
                    'total': total_titulo_publico,
                    'percentual': (total_titulo_publico / total_geral * 100) if total_geral > 0 else 0,
                    'percentual_rf': (total_titulo_publico / total_renda_fixa * 100) if total_renda_fixa > 0 else 0,
                    'detalhes': [{'tipo': k, 'valor': v, 'percentual': (v / total_titulo_publico * 100) if total_titulo_publico > 0 else 0}
                                for k, v in sorted(detalhes_publico.items(), key=lambda x: x[1], reverse=True)]
                },
                'total_geral': total_geral,
                'total_renda_fixa': total_renda_fixa
            }

            # === EXPOSIÇÃO RENDA FIXA (CUSTODIA_RF) ===
            # Buscar dados de custódia RF para o cliente selecionado
            try:
                # Buscar todos os dados de custódia RF do usuário
                res_custodia_rf = supabase.table("custodia_rf")\
                    .select("*")\
                    .eq("user_id", uid)\
                    .execute()

                current_app.logger.info(f"ASSET_ALLOCATION RF: Total de registros RF encontrados: {len(res_custodia_rf.data) if res_custodia_rf.data else 0}")

                if res_custodia_rf.data:
                    # Filtrar posições que pertencem ao cliente selecionado
                    # O cod_conta na custodia_rf deve corresponder ao codigo_xp ou codigo_mb do cliente

                    # DEBUG: Verificar tipos e valores
                    cod_conta_exemplo = res_custodia_rf.data[0].get('cod_conta') if res_custodia_rf.data else None
                    current_app.logger.info(f"ASSET_ALLOCATION RF DEBUG: cliente_selecionado='{cliente_selecionado}' (tipo: {type(cliente_selecionado).__name__})")
                    current_app.logger.info(f"ASSET_ALLOCATION RF DEBUG: cod_conta exemplo='{cod_conta_exemplo}' (tipo: {type(cod_conta_exemplo).__name__})")

                    posicoes_rf_cliente = [
                        pos for pos in res_custodia_rf.data
                        if str(pos.get('cod_conta')) == str(cliente_selecionado)
                    ]

                    current_app.logger.info(f"ASSET_ALLOCATION: Encontradas {len(posicoes_rf_cliente)} posições RF para cliente {cliente_selecionado}")

                    # Classificar por indexador
                    total_pos_fixado = 0.0
                    total_pre_fixado = 0.0
                    total_inflacao = 0.0
                    total_internacional = 0.0
                    total_rf = 0.0

                    # Acumuladores para taxas ponderadas
                    taxas_ponderadas = {}  # {indexador: {'total_ponderado': 0.0, 'total_custodia': 0.0}}

                    for pos_rf in posicoes_rf_cliente:
                        indexador = (pos_rf.get('indexador', '') or '').strip().upper()
                        custodia = _to_float(pos_rf.get('custodia', 0))
                        taxa_cliente = _to_float(pos_rf.get('taxa_cliente', 0))
                        total_rf += custodia

                        current_app.logger.debug(f"ASSET_ALLOCATION RF: indexador='{indexador}' | custodia={custodia} | taxa_cliente={taxa_cliente}")

                        # Classificação segundo as regras:
                        # Pós-fixado: % CDI, CDI +, LFT, Renda+, Selic
                        # Inflação: IPCA
                        # Prefixado: PRE, PRÉ
                        # Internacional: DOLAR PTAX

                        if indexador in ['% CDI', 'CDI +', 'LFT', 'RENDA+', 'SELIC']:
                            total_pos_fixado += custodia
                        elif indexador == 'IPCA':
                            total_inflacao += custodia
                        elif indexador in ['PRE', 'PRÉ']:
                            total_pre_fixado += custodia
                        elif indexador == 'DOLAR PTAX':
                            total_internacional += custodia

                        # Calcular taxa ponderada por indexador
                        # Normalizar indexador para agrupamento
                        indexador_normalizado = indexador
                        taxa_percentual_cdi = 0.0

                        # LFT e SELIC = 100% CDI
                        if indexador in ['LFT', 'SELIC']:
                            indexador_normalizado = '% CDI'
                            taxa_percentual_cdi = 100.0
                        # % CDI já está em percentual
                        elif indexador == '% CDI':
                            taxa_percentual_cdi = taxa_cliente
                        # CDI + precisa ser convertido para % CDI
                        # Fórmula: (100 + taxa_cdi_plus) / 100 * 100 = 100 + taxa_cdi_plus
                        elif indexador == 'CDI +':
                            indexador_normalizado = '% CDI'
                            taxa_percentual_cdi = 100.0 + taxa_cliente
                        # RENDA+ também considera 100% CDI
                        elif indexador == 'RENDA+':
                            indexador_normalizado = '% CDI'
                            taxa_percentual_cdi = 100.0
                        # Para outros indexadores (IPCA, PRE, DOLAR PTAX), usar taxa_cliente diretamente
                        else:
                            taxa_percentual_cdi = taxa_cliente

                        # Acumular para cálculo de média ponderada
                        if indexador_normalizado not in taxas_ponderadas:
                            taxas_ponderadas[indexador_normalizado] = {'total_ponderado': 0.0, 'total_custodia': 0.0}

                        taxas_ponderadas[indexador_normalizado]['total_ponderado'] += custodia * taxa_percentual_cdi
                        taxas_ponderadas[indexador_normalizado]['total_custodia'] += custodia

                    current_app.logger.info(f"ASSET_ALLOCATION RF: Totais - Pós: R$ {total_pos_fixado:,.2f} | Pré: R$ {total_pre_fixado:,.2f} | Inflação: R$ {total_inflacao:,.2f} | Internacional: R$ {total_internacional:,.2f} | Total RF: R$ {total_rf:,.2f}")

                    # Calcular taxas médias ponderadas
                    taxas_medias_ponderadas = {}
                    for idx, dados in taxas_ponderadas.items():
                        if dados['total_custodia'] > 0:
                            taxa_media = dados['total_ponderado'] / dados['total_custodia']
                            taxas_medias_ponderadas[idx] = taxa_media
                            current_app.logger.info(f"ASSET_ALLOCATION RF: Taxa média ponderada {idx}: {taxa_media:.2f}%")
                        else:
                            taxas_medias_ponderadas[idx] = 0.0

                    if total_rf > 0:
                        exposicao_rf = {
                            'pos_fixado': {
                                'total': total_pos_fixado,
                                'percentual': (total_pos_fixado / total_rf * 100) if total_rf > 0 else 0,
                                'taxa_ponderada': taxas_medias_ponderadas.get('% CDI', 0.0)
                            },
                            'pre_fixado': {
                                'total': total_pre_fixado,
                                'percentual': (total_pre_fixado / total_rf * 100) if total_rf > 0 else 0,
                                'taxa_ponderada': taxas_medias_ponderadas.get('PRE', 0.0) or taxas_medias_ponderadas.get('PRÉ', 0.0)
                            },
                            'inflacao': {
                                'total': total_inflacao,
                                'percentual': (total_inflacao / total_rf * 100) if total_rf > 0 else 0,
                                'taxa_ponderada': taxas_medias_ponderadas.get('IPCA', 0.0)
                            },
                            'internacional': {
                                'total': total_internacional,
                                'percentual': (total_internacional / total_rf * 100) if total_rf > 0 else 0,
                                'taxa_ponderada': taxas_medias_ponderadas.get('DOLAR PTAX', 0.0)
                            },
                            'total_rf': total_rf
                    }

            except Exception as e:
                current_app.logger.warning(f"Erro ao buscar dados de custódia RF: {e}")

            # === MAPA DE LIQUIDEZ ===
            try:
                # Buscar todos os CNPJs de fundos do cliente no diversificador
                res_fundos_cliente = supabase.table("diversificador")\
                    .select("cnpj_fundo, net")\
                    .eq("user_id", uid)\
                    .eq("cliente", cliente_selecionado)\
                    .not_.is_("cnpj_fundo", "null")\
                    .execute()

                if res_fundos_cliente.data:
                    # Agrupar por CNPJ e somar valores
                    fundos_agrupados = {}
                    for item in res_fundos_cliente.data:
                        cnpj_str = (item.get('cnpj_fundo') or '').strip()
                        net = _to_float(item.get('net', 0))
                        if cnpj_str:
                            # Converter CNPJ para inteiro para match com a tabela mapa_liquidez
                            try:
                                cnpj_int = int(cnpj_str.replace('.', '').replace('/', '').replace('-', ''))
                                fundos_agrupados[cnpj_int] = fundos_agrupados.get(cnpj_int, 0.0) + net
                            except (ValueError, AttributeError):
                                current_app.logger.warning(f"MAPA_LIQUIDEZ: CNPJ inválido no diversificador: {cnpj_str}")

                    current_app.logger.info(f"MAPA_LIQUIDEZ: Cliente possui {len(fundos_agrupados)} fundos únicos (CNPJs: {list(fundos_agrupados.keys())})")

                    # Buscar informações dos fundos na tabela mapa_liquidez
                    cnpjs_list = list(fundos_agrupados.keys())
                    if cnpjs_list:
                        # Tentar buscar como INTEGER primeiro
                        res_mapa = supabase.table("mapa_liquidez")\
                            .select("*")\
                            .in_("CNPJ_FUNDO", cnpjs_list)\
                            .execute()

                        current_app.logger.info(f"MAPA_LIQUIDEZ: Query com CNPJs como INT {cnpjs_list[:3]}... retornou {len(res_mapa.data) if res_mapa.data else 0} fundos")

                        # Se não encontrou nada, tentar como STRING
                        if not res_mapa.data:
                            cnpjs_str_list = [str(cnpj) for cnpj in cnpjs_list]
                            current_app.logger.info(f"MAPA_LIQUIDEZ: Tentando buscar como STRING: {cnpjs_str_list[:3]}...")
                            res_mapa = supabase.table("mapa_liquidez")\
                                .select("*")\
                                .in_("CNPJ_FUNDO", cnpjs_str_list)\
                                .execute()
                            current_app.logger.info(f"MAPA_LIQUIDEZ: Query com CNPJs como STRING retornou {len(res_mapa.data) if res_mapa.data else 0} fundos")

                        if res_mapa.data:
                            for fundo in res_mapa.data:
                                cnpj = fundo.get('CNPJ_FUNDO', '')
                                # Converter CNPJ para int para buscar no dicionário fundos_agrupados
                                try:
                                    cnpj_int = int(str(cnpj).replace('.', '').replace('/', '').replace('-', ''))
                                    valor_fundo = fundos_agrupados.get(cnpj_int, 0.0)
                                except (ValueError, AttributeError):
                                    current_app.logger.warning(f"MAPA_LIQUIDEZ: Erro ao converter CNPJ {cnpj} para int")
                                    valor_fundo = 0.0

                                mapa_liquidez.append({
                                    'nome_fundo': fundo.get('NOME_FUNDO', 'N/A'),
                                    'cnpj_fundo': str(cnpj),  # Converter para string para exibição
                                    'classificacao_cvm': fundo.get('CLASSIFICAÇÃO_CVM', 'N/A'),
                                    'resgate_total': fundo.get('RESGATE TOTAL', 0),
                                    'valor': valor_fundo
                                })
                                current_app.logger.info(f"MAPA_LIQUIDEZ: Adicionado fundo {fundo.get('NOME_FUNDO', '')[:30]} (CNPJ {cnpj}) com valor R$ {valor_fundo:.2f}")

                            current_app.logger.info(f"MAPA_LIQUIDEZ: Encontrados {len(mapa_liquidez)} fundos no mapa de liquidez")
                        else:
                            current_app.logger.warning(f"MAPA_LIQUIDEZ: Nenhum fundo encontrado na tabela mapa_liquidez para os CNPJs do cliente")
                else:
                    current_app.logger.info(f"MAPA_LIQUIDEZ: Cliente não possui fundos com CNPJ informado")

                # Adicionar valores de LFT da tabela custodia_rf
                try:
                    res_lft = supabase.table("custodia_rf")\
                        .select("nome_papel, custodia")\
                        .eq("user_id", uid)\
                        .eq("cod_conta", str(cliente_selecionado))\
                        .execute()

                    if res_lft.data:
                        total_lft = 0.0
                        for item in res_lft.data:
                            nome_papel = (item.get('nome_papel', '') or '').strip().upper()
                            if nome_papel.startswith('LFT'):
                                custodia = _to_float(item.get('custodia', 0))
                                total_lft += custodia

                        if total_lft > 0:
                            # Adicionar LFT ao mapa de liquidez
                            mapa_liquidez.append({
                                'nome_fundo': 'Tesouro Selic (LFT)',
                                'cnpj_fundo': 'N/A',
                                'classificacao_cvm': 'Renda Fixa',
                                'resgate_total': 0,  # D+0 (liquidez imediata)
                                'valor': total_lft
                            })
                            current_app.logger.info(f"MAPA_LIQUIDEZ: Adicionado LFT com total de R$ {total_lft:.2f}")

                except Exception as e:
                    current_app.logger.warning(f"MAPA_LIQUIDEZ: Erro ao buscar LFT da custodia_rf: {e}")

                # Ordenar por valor decrescente
                mapa_liquidez.sort(key=lambda x: x['valor'], reverse=True)

            except Exception as e:
                current_app.logger.warning(f"MAPA_LIQUIDEZ: Erro ao buscar mapa de liquidez: {e}")
                import traceback
                current_app.logger.warning(f"MAPA_LIQUIDEZ: Traceback: {traceback.format_exc()}")

        return render_template('clientes/asset_allocation.html',
                             clientes=clientes_com_nome,
                             letras_disponiveis=letras_disponiveis,
                             cliente_selecionado=cliente_selecionado,
                             cliente_selecionado_nome=cliente_selecionado_nome,
                             por_produto=por_produto,
                             por_sub_produto=por_sub_produto,
                             por_ativo=por_ativo,
                             por_emissor=por_emissor,
                             exposicao_por_tipo=exposicao_por_tipo,
                             credito_privado_vencimentos=credito_privado_vencimentos if cliente_selecionado else [],
                             credito_privado_emissores=credito_privado_emissores if cliente_selecionado else [],
                             exposicao_rf=exposicao_rf,
                             mapa_liquidez=mapa_liquidez)

    except Exception:
        current_app.logger.exception("Falha ao carregar Asset Allocation")
        flash("Falha ao carregar dados de Asset Allocation.", "warning")
        return render_template('clientes/asset_allocation.html',
                             clientes=[],
                             cliente_selecionado=None,
                             por_produto=[],
                             por_sub_produto=[],
                             por_ativo=[],
                             por_emissor=[],
                             credito_privado_vencimentos=[],
                             credito_privado_emissores=[],
                             exposicao_rf={})
